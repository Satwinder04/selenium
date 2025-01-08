import openpyxl
from openpyxl.styles import PatternFill




def getRowCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return(sheet.max_row)

def getColumnCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return(sheet.max_column)

def readData(file,sheetName,rownum,columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(rownum,columnno).value

def writeData(file,sheetName,rownum,columnno,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(rownum, columnno).value = data
    workbook.save(file)

def fillGreenColor(file,sheetName,rownum,columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    greenFill = PatternFill(start_color='60b212',
                       end_color='60b212',
                       fill_type='solid')

    sheet.cell(rownum,columnno).fill=greenFill
    workbook.save(file)


def fillRedColor(file,sheetName,rownum,columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    redFill = PatternFill(start_color='ff0000',
                       end_color='ff0000',
                       fill_type='solid')
    sheet.cell(rownum,columnno).fill=redFill
    workbook.save(file)




from selenium import webdriver
from selenium.webdriver.common.by import By

driver=webdriver.Chrome()
driver.get("https://www.moneycontrol.com/fixed-income/calculator/state-bank-of-india-sbi/fixed-deposit-calculator-SBI-BSB001.html")

f="dataDriven/intrest/simple_interest_data.xlsx"

rows=getRowCount(f,"sheet1")
print(rows)

import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

file = 'dataDriven/intrest/simple_interest_data.xlsx'
workbook = openpyxl.load_workbook(file)
sheet = workbook.active
deposit_amount = sheet.cell(2, 1).value
tenure = sheet.cell(2, 3).value
interest_rate = sheet.cell(2, 2).value

chrome_options = Options()
driver = webdriver.Chrome(options=chrome_options)
driver.get("https://www.moneycontrol.com/fixed-income/calculator/state-bank-of-india-sbi/fixed-deposit-calculator-SBI-BSB001.html")

wait = WebDriverWait(driver, 10)
deposit_field = wait.until(EC.presence_of_element_located((By.XPATH, "//*[@id='principal']")))
deposit_field.clear()
deposit_field.send_keys(str(deposit_amount))

time.sleep(2)
rate_field = driver.find_element(By.ID, "interest")
rate_field.clear()
rate_field.send_keys(str(interest_rate))

time.sleep(2)
tenure_field = driver.find_element(By.ID, "tenure")
tenure_field.clear()
tenure_field.send_keys(str(tenure))

time.sleep(2)
calculate_button = driver.find_element(By.XPATH, "//*[@id='fdMatVal']/div[2]/a[1]/img")
calculate_button.click()

time.sleep(3)
driver.quit()

def calculateInterest(deposit_amount, tenure, interest_rate):
    # Example calculation for simple interest
    return (deposit_amount * interest_rate * tenure) / 100  # Adjust this logic as needed

def processAllRows(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    max_row = sheet.max_row

    for row in range(2, max_row + 1):  # Start from row 2 to skip headers
        deposit_amount = sheet.cell(row, 1).value
        tenure = sheet.cell(row, 3).value
        interest_rate = sheet.cell(row, 2).value
        
        # Perform the calculation
        result = calculateInterest(deposit_amount, tenure, interest_rate)
        
        # Write the result back to the sheet (assuming you want to write it in column 4)
        writeData(file, sheetName, row, 4, result)

        # Fill color based on some condition
        if result > 1000:  # Example condition for green fill
            fillGreenColor(file, sheetName, row, 4)
        else:  # Example condition for red fill
            fillRedColor(file, sheetName, row, 4)

# Call the function to process all rows
processAllRows(f, "sheet1")
