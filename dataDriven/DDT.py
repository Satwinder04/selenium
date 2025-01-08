# # import os
# import openpyxl
#
#
# file=r"dataDriven/test.xlsx"
# workbook=openpyxl.load_workbook(file)
# sheet=workbook["Sheet1"]
# rows=sheet.max_row
# print(rows)
# cols=sheet.max_column
# print(cols)
#
# for r in range(1,rows+1):
#     for c in range(1,cols+1):
#         print(sheet.cell(r,c).value,end="        ")
#     print()

#
import openpyxl

file=r"dataDriven/test1.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook.active
# sheet=workbook["Sheet1"]
for r in range(1,6):
    for c in range(1,6):
        sheet.cell(r,c).value="wellcom"

workbook.save(file)


