import time
from datetime import datetime

from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import load_workbook

workbook=load_workbook('login_data.xlsx')
sheet=workbook.active
driver=webdriver.Chrome()
driver.maximize_window()
c=1
for row in sheet.iter_rows(min_row=2,max_row=sheet.max_row,values_only=True):
    username=row[0]
    password=row[1]
    expected=row[2]
    driver.get("https://online.btes.co.in/login/index.php")
    time.sleep(2)
    driver.find_element(By.XPATH, "//*[@id='username']").send_keys(username)
    time.sleep(1)
    driver.find_element(By.XPATH, "//*[@id='password']").send_keys(password)
    time.sleep(1)
    driver.find_element(By.XPATH, "//*[@id='loginbtn']").click()
    time.sleep(1)

    if expected == "Satwinder Singh":
        act = driver.find_element(By.XPATH, "//*[@class='usertext mr-1']").text
    else:
        act = driver.find_element(By.XPATH, "//a[text()='Invalid login, please try again']").text
    assert expected == act, f"test case {c} fail"
    c+=1
    time.sleep(1)
    timestamp=datetime.now().strftime("%Y%m%d_%h%M%S")
    
driver.quit()
