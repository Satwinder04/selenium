import time

from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import load_workbook

workbook=load_workbook('dataDriven/data.xlsx')
sheet=workbook.active
driver=webdriver.Chrome()
driver.maximize_window()

for row in sheet.iter_rows(min_row=2,max_row=sheet.max_row,values_only=True):
    username=row[0]
    password=row[1]
    driver.get("https://www.saucedemo.com/v1/")
    time.sleep(5)
    u=driver.find_element(By.ID,"user-name")
    u.send_keys(username)
    p=driver.find_element(By.ID,"password")
    p.send_keys(password)
    driver.find_element(By.ID,"login-button").click()
    # errormsg=driver.find_element(By.XPATH,"//h3[@data-test='error']")
    if username!="locked_out_user":
        time.sleep(2)
        driver.find_element(By.XPATH,"//button[normalize-space()='Open Menu']").click()
        time.sleep(2)
        driver.find_element(By.XPATH,"//a[@id='logout_sidebar_link']").click()
        time.sleep(2)

driver.quit()
